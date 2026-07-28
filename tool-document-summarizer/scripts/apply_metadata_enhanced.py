"""Preview metadata changes; write only with --apply and a backup directory."""

from __future__ import annotations

import argparse
import hashlib
import json
import shutil
import sys
import tempfile
from pathlib import Path


def load_inputs(summaries_path, mapping_path):
    summaries = json.loads(summaries_path.read_text(encoding="utf-8"))
    mapping = json.loads(mapping_path.read_text(encoding="utf-8"))
    if not isinstance(summaries, list) or not isinstance(mapping, dict):
        raise ValueError("summaries must be an array and mapping must be an object")
    seen = set()
    proposed = []
    for index, item in enumerate(summaries):
        if not isinstance(item, dict):
            raise ValueError(f"summary[{index}] must be an object")
        doc_id = item.get("id")
        summary = item.get("summary")
        tags = item.get("tags", [])
        if not isinstance(doc_id, str) or not doc_id or doc_id in seen:
            raise ValueError(f"summary[{index}] has a missing or duplicate id")
        seen.add(doc_id)
        if (
            not isinstance(summary, str)
            or not isinstance(tags, list)
            or any(not isinstance(tag, str) for tag in tags)
        ):
            raise ValueError(f"summary[{index}] has invalid summary or tags")
        source = mapping.get(doc_id)
        if not isinstance(source, str):
            raise ValueError(f"mapping is missing source id: {doc_id}")
        path = Path(source)
        if not path.is_file():
            raise ValueError(f"source file not found: {path}")
        proposed.append(
            {"id": doc_id, "path": path, "summary": summary, "tags": tags}
        )
    return proposed


def has_metadata(path):
    suffix = path.suffix.lower()
    if suffix == ".pdf":
        from pypdf import PdfReader
        metadata = PdfReader(str(path)).metadata or {}
        return bool(metadata.get("/Subject") or metadata.get("/Keywords"))
    if suffix == ".docx":
        from docx import Document
        props = Document(str(path)).core_properties
        return bool(props.subject or props.keywords)
    if suffix == ".pptx":
        from pptx import Presentation
        props = Presentation(str(path)).core_properties
        return bool(props.subject or props.keywords)
    if suffix in {".xlsx", ".xlsm"}:
        from openpyxl import load_workbook
        workbook = load_workbook(str(path), read_only=True)
        value = bool(workbook.properties.subject or workbook.properties.keywords)
        workbook.close()
        return value
    raise ValueError(f"unsupported source type: {suffix}")


def write_metadata(source, summary, tags):
    suffix = source.suffix.lower()
    temp_handle = tempfile.NamedTemporaryFile(
        delete=False, suffix=suffix, dir=str(source.parent)
    )
    temp_path = Path(temp_handle.name)
    temp_handle.close()
    try:
        if suffix == ".pdf":
            from pypdf import PdfReader, PdfWriter
            reader, writer = PdfReader(str(source)), PdfWriter()
            for page in reader.pages:
                writer.add_page(page)
            metadata = dict(reader.metadata or {})
            metadata.update({"/Subject": summary, "/Keywords": ", ".join(tags)})
            writer.add_metadata(metadata)
            with temp_path.open("wb") as handle:
                writer.write(handle)
        elif suffix == ".docx":
            from docx import Document
            document = Document(str(source))
            document.core_properties.subject = summary
            document.core_properties.keywords = ", ".join(tags)
            document.save(str(temp_path))
        elif suffix == ".pptx":
            from pptx import Presentation
            presentation = Presentation(str(source))
            presentation.core_properties.subject = summary
            presentation.core_properties.keywords = ", ".join(tags)
            presentation.save(str(temp_path))
        elif suffix in {".xlsx", ".xlsm"}:
            from openpyxl import load_workbook
            workbook = load_workbook(str(source), keep_vba=suffix == ".xlsm")
            workbook.properties.subject = summary
            workbook.properties.keywords = ", ".join(tags)
            workbook.save(str(temp_path))
            workbook.close()
        else:
            raise ValueError(f"unsupported source type: {suffix}")
        temp_path.replace(source)
    except Exception:
        temp_path.unlink(missing_ok=True)
        raise


def main() -> int:
    parser = argparse.ArgumentParser(description="Preview or apply document metadata")
    parser.add_argument("--summaries", required=True, type=Path)
    parser.add_argument("--mapping", required=True, type=Path)
    parser.add_argument("--backup-dir", type=Path)
    parser.add_argument("--apply", action="store_true")
    parser.add_argument("--overwrite-existing", action="store_true")
    args = parser.parse_args()
    if args.apply and not args.backup_dir:
        print("ERROR[authorization]: --apply requires --backup-dir", file=sys.stderr)
        return 2

    try:
        proposed = load_inputs(args.summaries, args.mapping)
        preview = [
            {
                "id": item["id"],
                "path": str(item["path"].resolve()),
                "subject": item["summary"],
                "keywords": item["tags"],
            }
            for item in proposed
        ]
        print(json.dumps({"mode": "apply" if args.apply else "preview", "changes": preview}, ensure_ascii=False, indent=2))
        if not args.apply:
            print("PREVIEW_ONLY: add --apply --backup-dir <dir> after review")
            return 0

        args.backup_dir.mkdir(parents=True, exist_ok=True)
        for item in proposed:
            if has_metadata(item["path"]) and not args.overwrite_existing:
                raise ValueError(
                    f"source already has metadata: {item['path']}; use --overwrite-existing explicitly"
                )
            digest = hashlib.sha256(str(item["path"].resolve()).encode("utf-8")).hexdigest()[:12]
            backup = args.backup_dir / f"{digest}_{item['path'].name}"
            if backup.exists():
                raise ValueError(f"backup already exists: {backup}")
            shutil.copy2(item["path"], backup)
            write_metadata(item["path"], item["summary"], item["tags"])
        print(f"APPLY_PASS: {len(proposed)} source files changed; backups: {args.backup_dir.resolve()}")
    except (OSError, UnicodeError, json.JSONDecodeError, ValueError, ImportError) as exc:
        print(f"ERROR[metadata]: {exc}", file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
