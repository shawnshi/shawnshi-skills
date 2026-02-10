#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
@Input:  document_summaries_enhanced.json, file_id_mapping.json
@Output: Persistent file properties (Subject/Keywords) updated in original documents
@Pos:    Adapter Layer. Synchronizes semantic intelligence back to the physical file system.

!!! Maintenance Protocol: Ensure Excel fallback mechanism is tested against locked files.
!!! Supports Subject (Summaries) and Keywords (Tags).

优化版元数据应用器
- 解决Excel兼容性问题（三层回退机制）
- 增量处理（跳过已有元数据的文件）
- 详细日志记录
- 并行处理提升性能
"""
import json
import sys
import io
import logging
import argparse
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime

# 设置标准输出编码为 UTF-8
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

try:
    from pypdf import PdfReader, PdfWriter
except ImportError:
    PdfReader = PdfWriter = None

try:
    from docx import Document
except ImportError:
    Document = None

try:
    from pptx import Presentation
except ImportError:
    Presentation = None

try:
    import openpyxl
except ImportError:
    openpyxl = None


# 配置日志
def setup_logging(log_dir, log_level=logging.INFO):
    """设置日志系统"""
    log_file = Path(log_dir) / f"metadata_application_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
    logging.basicConfig(
        level=log_level,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_file, encoding='utf-8'),
            logging.StreamHandler(sys.stdout)
        ]
    )
    return log_file


def has_metadata(file_path):
    """检查文件是否已有元数据"""
    ext = file_path.suffix.lower()
    try:
        if ext == '.pdf' and PdfReader:
            reader = PdfReader(str(file_path))
            metadata = reader.metadata
            if metadata:
                subject = metadata.get('/Subject', '')
                keywords = metadata.get('/Keywords', '')
                comments = metadata.get('/Comments', '')
                return bool(subject and keywords and comments)

        elif ext == '.docx' and Document:
            doc = Document(str(file_path))
            props = doc.core_properties
            return bool(props.subject and props.keywords and props.comments)

        elif ext == '.pptx' and Presentation:
            prs = Presentation(str(file_path))
            props = prs.core_properties
            return bool(props.subject and props.keywords and props.comments)

        elif ext == '.xlsx' and openpyxl:
            wb = openpyxl.load_workbook(str(file_path), read_only=True)
            props = wb.properties
            return bool(props.subject and props.keywords and props.description)

    except Exception as e:
        logging.warning(f"检查元数据失败 {file_path.name}: {e}")

    return False


def apply_pdf_metadata(file_path, summary, tags):
    """为 PDF 文件应用元数据"""
    if not PdfReader or not PdfWriter:
        return False, "PDF 库未安装"

    try:
        reader = PdfReader(str(file_path))
        writer = PdfWriter()

        # 复制所有页面
        for page in reader.pages:
            writer.add_page(page)

        # 设置元数据
        writer.add_metadata({
            '/Subject': summary,
            '/Keywords': ', '.join(tags),
            '/Comments': summary
        })

        # 保存
        with open(str(file_path), 'wb') as f:
            writer.write(f)

        return True, "成功"
    except Exception as e:
        return False, str(e)


def apply_docx_metadata(file_path, summary, tags):
    """为 DOCX 文件应用元数据"""
    if not Document:
        return False, "DOCX 库未安装"

    try:
        doc = Document(str(file_path))
        core_props = doc.core_properties

        core_props.subject = summary
        core_props.keywords = ', '.join(tags)
        core_props.comments = summary

        doc.save(str(file_path))
        return True, "成功"
    except Exception as e:
        return False, str(e)


def apply_pptx_metadata(file_path, summary, tags):
    """为 PPTX 文件应用元数据"""
    if not Presentation:
        return False, "PPTX 库未安装"

    try:
        prs = Presentation(str(file_path))
        core_props = prs.core_properties

        core_props.subject = summary
        core_props.keywords = ', '.join(tags)
        core_props.comments = summary

        prs.save(str(file_path))
        return True, "成功"
    except Exception as e:
        return False, str(e)


def apply_xlsx_metadata_safe(file_path, summary, tags):
    """为 XLSX 文件应用元数据（增强兼容性 - 三层回退）"""
    if not openpyxl:
        return False, "XLSX 库未安装"

    # 方法1：标准方式
    try:
        wb = openpyxl.load_workbook(str(file_path))
        props = wb.properties

        props.subject = summary
        props.keywords = ', '.join(tags)
        props.description = summary  # Excel 属性中 description 通常对应备注

        wb.save(str(file_path))
        return True, "成功"
    except Exception as e1:
        # 方法2：只读模式加载，然后写入
        try:
            wb = openpyxl.load_workbook(str(file_path), data_only=True)
            props = wb.properties

            props.subject = summary
            props.keywords = ', '.join(tags)
            props.description = summary

            wb.save(str(file_path))
            return True, "成功（兼容模式）"
        except Exception as e2:
            # 方法3：创建临时文件
            try:
                import tempfile
                import shutil

                # 创建临时文件
                temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
                temp_path = temp_file.name
                temp_file.close()

                # 复制原文件
                shutil.copy2(str(file_path), temp_path)

                # 尝试修改临时文件
                wb = openpyxl.load_workbook(temp_path)
                props = wb.properties
                props.subject = summary
                props.keywords = ', '.join(tags)
                props.description = summary
                wb.save(temp_path)

                # 替换原文件
                shutil.move(temp_path, str(file_path))
                return True, "成功（临时文件模式）"
            except Exception as e3:
                logging.error(f"所有方法均失败: {e1} | {e2} | {e3}")
                return False, f"兼容性问题: {str(e1)[:50]}"


def process_single_file(file_info, file_mapping, skip_existing=True):
    """处理单个文件"""
    file_id = file_info['id']
    summary = file_info['summary']
    tags = file_info['tags']
    filename = file_info['filename']

    # 获取文件路径
    if file_id not in file_mapping:
        return {
            'filename': filename,
            'status': 'skip',
            'reason': '文件ID未找到'
        }

    file_path = Path(file_mapping[file_id])

    if not file_path.exists():
        return {
            'filename': filename,
            'status': 'skip',
            'reason': '文件不存在'
        }

    # 检查是否已有元数据（增量处理）
    if skip_existing and has_metadata(file_path):
        return {
            'filename': filename,
            'status': 'skip',
            'reason': '已有元数据'
        }

    # 根据文件类型应用元数据
    ext = file_path.suffix.lower()

    if ext == '.pdf':
        success, msg = apply_pdf_metadata(file_path, summary, tags)
    elif ext == '.docx':
        success, msg = apply_docx_metadata(file_path, summary, tags)
    elif ext == '.pptx':
        success, msg = apply_pptx_metadata(file_path, summary, tags)
    elif ext == '.xlsx':
        success, msg = apply_xlsx_metadata_safe(file_path, summary, tags)
    else:
        return {
            'filename': filename,
            'status': 'skip',
            'reason': '不支持的文件类型'
        }

    if success:
        return {
            'filename': filename,
            'status': 'success',
            'reason': msg
        }
    else:
        return {
            'filename': filename,
            'status': 'fail',
            'reason': msg
        }


def apply_metadata_parallel(summaries, file_mapping, max_workers=5, skip_existing=True):
    """并行处理文档元数据"""
    results = {
        'success': [],
        'fail': [],
        'skip': []
    }

    print(f"\n开始并行处理 (工作线程: {max_workers})...\n")

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        # 提交所有任务
        futures = {
            executor.submit(process_single_file, file_info, file_mapping, skip_existing): file_info
            for file_info in summaries
        }

        # 处理完成的任务
        completed = 0
        total = len(summaries)

        for future in as_completed(futures):
            completed += 1
            file_info = futures[future]

            try:
                result = future.result()

                if result['status'] == 'success':
                    results['success'].append(result)
                    print(f"[{completed}/{total}] ✓ {result['filename']}")
                    logging.info(f"成功: {result['filename']} - {result['reason']}")
                elif result['status'] == 'fail':
                    results['fail'].append(result)
                    print(f"[{completed}/{total}] ✗ {result['filename']} ({result['reason'][:30]}...)")
                    logging.error(f"失败: {result['filename']} - {result['reason']}")
                else:
                    results['skip'].append(result)
                    if completed % 50 == 0:
                        print(f"[{completed}/{total}] ⊘ {result['filename']} ({result['reason']})")
                    logging.debug(f"跳过: {result['filename']} - {result['reason']}")

            except Exception as e:
                results['fail'].append({
                    'filename': file_info['filename'],
                    'status': 'fail',
                    'reason': f'异常: {str(e)}'
                })
                print(f"[{completed}/{total}] ✗ {file_info['filename']} (异常: {str(e)[:30]}...)")
                logging.exception(f"异常: {file_info['filename']}")

            # 每100个显示进度
            if completed % 100 == 0:
                print(f"\n--- 进度: {completed}/{total} ({completed*100//total}%) ---")
                print(f"    成功: {len(results['success'])} | 失败: {len(results['fail'])} | 跳过: {len(results['skip'])}\n")

    return results


def main():
    parser = argparse.ArgumentParser(description='优化版元数据应用器')
    parser.add_argument('--summaries', required=True, help='摘要JSON文件路径')
    parser.add_argument('--mapping', required=True, help='文件ID映射JSON路径')
    parser.add_argument('--workers', type=int, default=5, help='并行工作线程数 (默认: 5)')
    parser.add_argument('--force', action='store_true', help='强制处理所有文件，包括已有元数据的')
    parser.add_argument('--log-dir', default='.', help='日志文件保存目录')
    args = parser.parse_args()

    # 设置日志
    log_file = setup_logging(args.log_dir)
    print(f"日志文件: {log_file}\n")

    # 读取摘要数据
    print("正在读取摘要数据...")
    with open(args.summaries, "r", encoding="utf-8") as f:
        summaries = json.load(f)

    # 读取文件映射
    with open(args.mapping, "r", encoding="utf-8") as f:
        file_mapping = json.load(f)

    print(f"共有 {len(summaries)} 个文档需要应用元数据")

    # 增量处理模式
    skip_existing = not args.force
    print(f"增量处理模式: {'开启' if skip_existing else '关闭'} {'(将跳过已有元数据的文件)' if skip_existing else '(将处理所有文件)'}\n")

    # 并行应用元数据
    results = apply_metadata_parallel(
        summaries,
        file_mapping,
        max_workers=args.workers,
        skip_existing=skip_existing
    )

    # 显示总结
    print("\n" + "="*60)
    print("应用元数据完成！")
    print("="*60)
    print(f"✓ 成功: {len(results['success'])} 个文档")
    print(f"✗ 失败: {len(results['fail'])} 个文档")
    print(f"⊘ 跳过: {len(results['skip'])} 个文档")
    print(f"总计: {len(summaries)} 个文档")
    print("="*60)

    # 保存失败列表
    if results['fail']:
        fail_file = Path(args.log_dir) / "metadata_application_failures.json"
        with open(fail_file, "w", encoding="utf-8") as f:
            json.dump(results['fail'], f, ensure_ascii=False, indent=2)
        print(f"\n失败文件列表已保存到: {fail_file}")

    print(f"\n详细日志已保存到: {log_file}")

    # 返回退出码
    return 0 if len(results['fail']) == 0 else 1


if __name__ == "__main__":
    sys.exit(main())
